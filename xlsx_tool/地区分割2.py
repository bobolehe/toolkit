import openpyxl
import json

# 打开Excel文件
workbook = openpyxl.load_workbook('IPV6归属地-API（区县级）_20231017114753.xlsx')

# 获取工作表（假设这里是第一个工作表）
sheet = workbook.active

# 获取行数
max_row_num = sheet.max_row

for i in range(2, max_row_num + 1):
    row_list2 = []
    for row in sheet[i]:
        row_list2.append(row.value)
    dz = ""
    if row_list2[6] == row_list2[7]:
        dz = row_list2[6] + row_list2[8]
    else:
        dz = row_list2[6] + row_list2[7] + row_list2[8]
    sheet.cell(row=i, column=20, value=dz)

# 保存修改后的Excel文件
workbook.save('IPv6归属地-API（高精准-公安版）_20231017114728_updated.xlsx')

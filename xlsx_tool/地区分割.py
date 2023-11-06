import openpyxl
import json

# 打开Excel文件
workbook = openpyxl.load_workbook('IPv4归属地-API（高精准-公安版）_20231017114728.xlsx')

# 获取工作表（假设这里是第一个工作表）
sheet = workbook.active

# 选择要读取的列（例如列N）
column_to_read = 14
column_to_write = 18

# 使用 iter_cols 方法来获取指定列的数据
for col in sheet.iter_cols(min_col=column_to_read, max_col=column_to_read):
    for cell in col:
        if cell.row == 1:
            continue
        data = json.loads(cell.value)
        if not data:
            continue
        # dz = ""
        # if data[0]['prov'] == data[0]['city']:
        #     dz = data[0]['prov']
        # else:
        #     dz = data[0]['prov'] + data[0]['city'] + data[0]['district']
        # 获取列表中的第一个元素
        data_to_write = data[0]['prov']
        # 将数据写入指定行的指定列
        sheet.cell(row=cell.row, column=column_to_write, value=data_to_write)

# 保存修改后的Excel文件
workbook.save('IPv4归属地-API（高精准-公安版）_20231017114728_添加省级列.xlsx')

# # 输出列数据
# print(column_data)
